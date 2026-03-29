import csv
import io
import json
import math
import os
import tempfile
import webbrowser
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from typing import List, Tuple

try:
    from pyproj import CRS, Transformer
    HAS_PYPROJ = True
except Exception:
    CRS = None  # type: ignore
    Transformer = None  # type: ignore
    HAS_PYPROJ = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except Exception:
    openpyxl = None  # type: ignore
    HAS_OPENPYXL = False

try:
    import matplotlib.pyplot as plt
    HAS_MPL = True
except Exception:
    plt = None  # type: ignore
    HAS_MPL = False

import main as core


LANG = {
    "ru": {
        "title": "GeoMate v1 - Геодезическое приложение",
        "lang": "Язык / Language",
        "tab_crs": "Преобразование CRS",
        "tab_geo": "Геодезические задачи",
        "tab_special": "Спецрасчет (пп.1-9)",
        "input_epsg": "Входная CRS (EPSG или WKT):",
        "output_epsg": "Выходная CRS (EPSG или WKT):",
        "coord_order": "Порядок столбцов:",
        "xy": "X Y",
        "xyz": "X Y Z",
        "latlon": "Lat Lon",
        "latlonh": "Lat Lon H",
        "import_txt": "Импорт TXT",
        "transform": "Преобразовать",
        "export_txt": "Экспорт TXT",
        "import_table": "Импорт CSV/XLSX",
        "export_table": "Экспорт CSV/XLSX",
        "export_kml": "Экспорт KML",
        "export_gpx": "Экспорт GPX",
        "export_dxf": "Экспорт DXF",
        "preview_map": "Мини-карта",
        "export_pdf": "Схема PDF",
        "save_project": "Сохранить проект",
        "load_project": "Загрузить проект",
        "datum_preset": "Пресет датума:",
        "height_model": "Модель высот:",
        "geoid_n": "N геоида (м):",
        "apply_height": "Применять H = h - N",
        "pyproj_missing": "Модуль pyproj не найден. Установите: pip install pyproj",
        "xlsx_missing": "Для XLSX нужен openpyxl: pip install openpyxl",
        "mpl_missing": "Для карты/PDF нужен matplotlib: pip install matplotlib",
        "history": "История",
        "dms_mode": "Формат углов:",
        "deg_mode": "Десятичные градусы",
        "dms_mode_value": "Градусы/Минуты/Секунды",
        "run_reduction": "Редукция (k,S->S')",
        "run_misclosure": "Невязки хода",
        "dragdrop_info": "Drag-and-drop: если не работает в системе, используйте кнопки Импорт.",
        "input_data": "Входные данные (по строкам):",
        "output_data": "Результат:",
        "inverse_title": "Обратная задача (B1,L1 -> B2,L2)",
        "forward_title": "Прямая задача (B1,L1,A,S)",
        "run_inverse": "Решить обратную",
        "run_forward": "Решить прямую",
        "special_run": "Выполнить пп.1-9",
        "b1": "B1 (deg)",
        "l1": "L1 (deg)",
        "h1": "H1 (m)",
        "b2": "B2 (deg)",
        "l2": "L2 (deg)",
        "h2": "H2 (m)",
        "az": "A12 (deg)",
        "dist": "S12 (m)",
        "ok": "Готово.",
        "err_parse": "Ошибка разбора строки:",
    },
    "en": {
        "title": "GeoMate v1 - Geodetic Application",
        "lang": "Language / Язык",
        "tab_crs": "CRS Transform",
        "tab_geo": "Geodetic Tools",
        "tab_special": "Special workflow (items 1-9)",
        "input_epsg": "Input CRS (EPSG or WKT):",
        "output_epsg": "Output CRS (EPSG or WKT):",
        "coord_order": "Column order:",
        "xy": "X Y",
        "xyz": "X Y Z",
        "latlon": "Lat Lon",
        "latlonh": "Lat Lon H",
        "import_txt": "Import TXT",
        "transform": "Transform",
        "export_txt": "Export TXT",
        "import_table": "Import CSV/XLSX",
        "export_table": "Export CSV/XLSX",
        "export_kml": "Export KML",
        "export_gpx": "Export GPX",
        "export_dxf": "Export DXF",
        "preview_map": "Mini-map",
        "export_pdf": "PDF sketch",
        "save_project": "Save Project",
        "load_project": "Load Project",
        "datum_preset": "Datum preset:",
        "height_model": "Height model:",
        "geoid_n": "Geoid N (m):",
        "apply_height": "Apply H = h - N",
        "pyproj_missing": "pyproj module not found. Install: pip install pyproj",
        "xlsx_missing": "XLSX requires openpyxl: pip install openpyxl",
        "mpl_missing": "Map/PDF requires matplotlib: pip install matplotlib",
        "history": "History",
        "dms_mode": "Angle format:",
        "deg_mode": "Decimal degrees",
        "dms_mode_value": "Degrees/Minutes/Seconds",
        "run_reduction": "Reduction (k,S->S')",
        "run_misclosure": "Traverse misclosure",
        "dragdrop_info": "Drag-and-drop: if unsupported on your system, use Import buttons.",
        "input_data": "Input data (one point per line):",
        "output_data": "Output:",
        "inverse_title": "Inverse geodetic problem (B1,L1 -> B2,L2)",
        "forward_title": "Direct geodetic problem (B1,L1,A,S)",
        "run_inverse": "Run inverse",
        "run_forward": "Run forward",
        "special_run": "Run items 1-9",
        "b1": "B1 (deg)",
        "l1": "L1 (deg)",
        "h1": "H1 (m)",
        "b2": "B2 (deg)",
        "l2": "L2 (deg)",
        "h2": "H2 (m)",
        "az": "A12 (deg)",
        "dist": "S12 (m)",
        "ok": "Done.",
        "err_parse": "Line parse error:",
    },
}


def parse_row(line: str) -> List[float]:
    line = line.strip()
    if not line:
        return []
    for sep in (";", ",", "\t"):
        line = line.replace(sep, " ")
    parts = [p for p in line.split(" ") if p]
    return [float(p.replace(",", ".")) for p in parts]


def angle_from_user(s: str, use_dms: bool) -> float:
    s = s.strip().replace(",", ".")
    if not use_dms:
        return float(s)
    # DMS formats accepted: "55 30 10.2" or "55:30:10.2"
    s = s.replace(":", " ")
    parts = [p for p in s.split(" ") if p]
    if len(parts) == 1:
        return float(parts[0])
    deg = float(parts[0])
    mins = float(parts[1]) if len(parts) > 1 else 0.0
    secs = float(parts[2]) if len(parts) > 2 else 0.0
    sign = -1 if deg < 0 else 1
    return sign * (abs(deg) + mins / 60.0 + secs / 3600.0)


class GeoMateApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.lang = "ru"
        self.text = LANG[self.lang]
        self.root.title(self.text["title"])
        self.root.geometry("1200x760")
        self.root.minsize(1050, 700)

        style = ttk.Style(self.root)
        style.theme_use("clam")

        self.top = ttk.Frame(root, padding=8)
        self.top.pack(fill=tk.X)
        ttk.Label(self.top, text=self.text["lang"]).pack(side=tk.LEFT)
        self.lang_var = tk.StringVar(value="RU")
        self.lang_box = ttk.Combobox(self.top, textvariable=self.lang_var, values=["RU", "EN"], width=6, state="readonly")
        self.lang_box.pack(side=tk.LEFT, padx=6)
        self.lang_box.bind("<<ComboboxSelected>>", self.on_language_change)

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        self.tab_crs = ttk.Frame(self.notebook, padding=10)
        self.tab_geo = ttk.Frame(self.notebook, padding=10)
        self.tab_special = ttk.Frame(self.notebook, padding=10)

        self.notebook.add(self.tab_crs, text=self.text["tab_crs"])
        self.notebook.add(self.tab_geo, text=self.text["tab_geo"])
        self.notebook.add(self.tab_special, text=self.text["tab_special"])

        self.build_crs_tab()
        self.build_geo_tab()
        self.build_special_tab()
        self.history_lines: List[str] = []

    def log(self, text: str) -> None:
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{stamp}] {text}"
        self.history_lines.append(line)
        try:
            with open("geomate_history.log", "a", encoding="utf-8") as f:
                f.write(line + "\n")
        except Exception:
            pass
        if hasattr(self, "history_text"):
            self.history_text.insert("end", line + "\n")
            self.history_text.see("end")

    def on_language_change(self, _event=None) -> None:
        self.lang = "en" if self.lang_var.get() == "EN" else "ru"
        self.text = LANG[self.lang]
        self.root.title(self.text["title"])
        self.refresh_labels()

    def refresh_labels(self) -> None:
        self.notebook.tab(0, text=self.text["tab_crs"])
        self.notebook.tab(1, text=self.text["tab_geo"])
        self.notebook.tab(2, text=self.text["tab_special"])

        self.lbl_in.configure(text=self.text["input_epsg"])
        self.lbl_out.configure(text=self.text["output_epsg"])
        self.lbl_order.configure(text=self.text["coord_order"])
        self.btn_import.configure(text=self.text["import_txt"])
        self.btn_import_table.configure(text=self.text["import_table"])
        self.btn_transform.configure(text=self.text["transform"])
        self.btn_export.configure(text=self.text["export_txt"])
        self.btn_export_table.configure(text=self.text["export_table"])
        self.btn_export_kml.configure(text=self.text["export_kml"])
        self.btn_export_gpx.configure(text=self.text["export_gpx"])
        self.btn_export_dxf.configure(text=self.text["export_dxf"])
        self.btn_preview_map.configure(text=self.text["preview_map"])
        self.btn_save_project.configure(text=self.text["save_project"])
        self.btn_load_project.configure(text=self.text["load_project"])
        self.lbl_input_data.configure(text=self.text["input_data"])
        self.lbl_output_data.configure(text=self.text["output_data"])
        self.order_box.configure(values=[self.text["xy"], self.text["xyz"], self.text["latlon"], self.text["latlonh"]])
        self.lbl_datum.configure(text=self.text["datum_preset"])
        self.lbl_height_model.configure(text=self.text["height_model"])
        self.lbl_geoid_n.configure(text=self.text["geoid_n"])
        self.chk_apply_height.configure(text=self.text["apply_height"])
        self.lbl_dms_mode.configure(text=self.text["dms_mode"])
        self.angle_mode_box.configure(values=[self.text["deg_mode"], self.text["dms_mode_value"]])
        self.btn_reduction.configure(text=self.text["run_reduction"])
        self.btn_misclosure.configure(text=self.text["run_misclosure"])
        self.lbl_history.configure(text=self.text["history"])

        self.inv_title.configure(text=self.text["inverse_title"])
        self.fwd_title.configure(text=self.text["forward_title"])
        self.btn_inverse.configure(text=self.text["run_inverse"])
        self.btn_forward.configure(text=self.text["run_forward"])

        self.btn_special.configure(text=self.text["special_run"])
        for k, lbl in self.special_labels.items():
            lbl.configure(text=self.text[k])

    def build_crs_tab(self) -> None:
        top = ttk.Frame(self.tab_crs)
        top.pack(fill=tk.X)

        self.lbl_in = ttk.Label(top, text=self.text["input_epsg"])
        self.lbl_in.grid(row=0, column=0, sticky="w")
        self.entry_in_crs = ttk.Entry(top, width=34)
        self.entry_in_crs.insert(0, "EPSG:4326")
        self.entry_in_crs.grid(row=0, column=1, padx=8, pady=4, sticky="w")

        self.lbl_out = ttk.Label(top, text=self.text["output_epsg"])
        self.lbl_out.grid(row=0, column=2, sticky="w")
        self.entry_out_crs = ttk.Entry(top, width=34)
        self.entry_out_crs.insert(0, "EPSG:3857")
        self.entry_out_crs.grid(row=0, column=3, padx=8, pady=4, sticky="w")

        self.lbl_order = ttk.Label(top, text=self.text["coord_order"])
        self.lbl_order.grid(row=1, column=0, sticky="w")
        self.order_var = tk.StringVar(value=self.text["latlon"])
        self.order_box = ttk.Combobox(top, textvariable=self.order_var, values=[self.text["xy"], self.text["xyz"], self.text["latlon"], self.text["latlonh"]], width=12, state="readonly")
        self.order_box.grid(row=1, column=1, sticky="w", padx=8, pady=4)

        self.lbl_datum = ttk.Label(top, text=self.text["datum_preset"])
        self.lbl_datum.grid(row=1, column=2, sticky="w")
        self.datum_var = tk.StringVar(value="WGS84")
        self.datum_box = ttk.Combobox(
            top,
            textvariable=self.datum_var,
            values=["WGS84", "SK-42", "SK-95", "PZ-90.11", "Custom"],
            width=12,
            state="readonly",
        )
        self.datum_box.grid(row=1, column=3, sticky="w", padx=8, pady=4)

        self.lbl_height_model = ttk.Label(top, text=self.text["height_model"])
        self.lbl_height_model.grid(row=2, column=0, sticky="w")
        self.height_model_var = tk.StringVar(value="Ellipsoidal")
        self.height_model_box = ttk.Combobox(
            top,
            textvariable=self.height_model_var,
            values=["Ellipsoidal", "Orthometric (N)", "Custom"],
            width=18,
            state="readonly",
        )
        self.height_model_box.grid(row=2, column=1, sticky="w", padx=8, pady=4)

        self.lbl_geoid_n = ttk.Label(top, text=self.text["geoid_n"])
        self.lbl_geoid_n.grid(row=2, column=2, sticky="w")
        self.entry_geoid_n = ttk.Entry(top, width=12)
        self.entry_geoid_n.insert(0, "0.0")
        self.entry_geoid_n.grid(row=2, column=3, sticky="w", padx=8, pady=4)

        self.apply_height_var = tk.BooleanVar(value=False)
        self.chk_apply_height = ttk.Checkbutton(top, text=self.text["apply_height"], variable=self.apply_height_var)
        self.chk_apply_height.grid(row=3, column=0, columnspan=2, sticky="w", pady=4)

        buttons = ttk.Frame(self.tab_crs)
        buttons.pack(fill=tk.X, pady=6)
        self.btn_import = ttk.Button(buttons, text=self.text["import_txt"], command=self.import_txt)
        self.btn_import.pack(side=tk.LEFT, padx=2)
        self.btn_import_table = ttk.Button(buttons, text=self.text["import_table"], command=self.import_table)
        self.btn_import_table.pack(side=tk.LEFT, padx=2)
        self.btn_transform = ttk.Button(buttons, text=self.text["transform"], command=self.run_transform)
        self.btn_transform.pack(side=tk.LEFT, padx=2)
        self.btn_export = ttk.Button(buttons, text=self.text["export_txt"], command=self.export_txt)
        self.btn_export.pack(side=tk.LEFT, padx=2)
        self.btn_export_table = ttk.Button(buttons, text=self.text["export_table"], command=self.export_table)
        self.btn_export_table.pack(side=tk.LEFT, padx=2)
        self.btn_export_kml = ttk.Button(buttons, text=self.text["export_kml"], command=self.export_kml)
        self.btn_export_kml.pack(side=tk.LEFT, padx=2)
        self.btn_export_gpx = ttk.Button(buttons, text=self.text["export_gpx"], command=self.export_gpx)
        self.btn_export_gpx.pack(side=tk.LEFT, padx=2)
        self.btn_export_dxf = ttk.Button(buttons, text=self.text["export_dxf"], command=self.export_dxf)
        self.btn_export_dxf.pack(side=tk.LEFT, padx=2)
        self.btn_preview_map = ttk.Button(buttons, text=self.text["preview_map"], command=self.preview_map)
        self.btn_preview_map.pack(side=tk.LEFT, padx=2)
        self.btn_save_project = ttk.Button(buttons, text=self.text["save_project"], command=self.save_project)
        self.btn_save_project.pack(side=tk.LEFT, padx=2)
        self.btn_load_project = ttk.Button(buttons, text=self.text["load_project"], command=self.load_project)
        self.btn_load_project.pack(side=tk.LEFT, padx=2)

        if not HAS_PYPROJ:
            self.btn_transform.state(["disabled"])
            self.btn_import_table.state(["disabled"])
            self.btn_export_table.state(["disabled"])

        io_frame = ttk.Panedwindow(self.tab_crs, orient=tk.HORIZONTAL)
        io_frame.pack(fill=tk.BOTH, expand=True)

        left = ttk.Frame(io_frame)
        right = ttk.Frame(io_frame)
        io_frame.add(left, weight=1)
        io_frame.add(right, weight=1)

        self.lbl_input_data = ttk.Label(left, text=self.text["input_data"])
        self.lbl_input_data.pack(anchor="w")
        self.input_text = tk.Text(left, height=30, wrap="none")
        self.input_text.pack(fill=tk.BOTH, expand=True)

        self.lbl_output_data = ttk.Label(right, text=self.text["output_data"])
        self.lbl_output_data.pack(anchor="w")
        self.output_text = tk.Text(right, height=30, wrap="none")
        self.output_text.pack(fill=tk.BOTH, expand=True)

        ttk.Label(self.tab_crs, text=self.text["dragdrop_info"]).pack(anchor="w", pady=4)

    def build_geo_tab(self) -> None:
        frame = ttk.Frame(self.tab_geo)
        frame.pack(fill=tk.BOTH, expand=True)

        self.lbl_dms_mode = ttk.Label(frame, text=self.text["dms_mode"])
        self.lbl_dms_mode.grid(row=0, column=0, sticky="w")
        self.angle_mode_var = tk.StringVar(value=self.text["deg_mode"])
        self.angle_mode_box = ttk.Combobox(
            frame,
            textvariable=self.angle_mode_var,
            values=[self.text["deg_mode"], self.text["dms_mode_value"]],
            width=24,
            state="readonly",
        )
        self.angle_mode_box.grid(row=0, column=1, sticky="w", padx=6)

        self.inv_title = ttk.Label(frame, text=self.text["inverse_title"], font=("Segoe UI", 10, "bold"))
        self.inv_title.grid(row=1, column=0, columnspan=4, sticky="w", pady=(8, 6))

        ttk.Label(frame, text=self.text["b1"]).grid(row=2, column=0, sticky="w")
        self.inv_b1 = ttk.Entry(frame, width=12)
        self.inv_b1.grid(row=2, column=1, padx=6, pady=2)
        ttk.Label(frame, text=self.text["l1"]).grid(row=2, column=2, sticky="w")
        self.inv_l1 = ttk.Entry(frame, width=12)
        self.inv_l1.grid(row=2, column=3, padx=6, pady=2)
        ttk.Label(frame, text=self.text["b2"]).grid(row=3, column=0, sticky="w")
        self.inv_b2 = ttk.Entry(frame, width=12)
        self.inv_b2.grid(row=3, column=1, padx=6, pady=2)
        ttk.Label(frame, text=self.text["l2"]).grid(row=3, column=2, sticky="w")
        self.inv_l2 = ttk.Entry(frame, width=12)
        self.inv_l2.grid(row=3, column=3, padx=6, pady=2)

        self.btn_inverse = ttk.Button(frame, text=self.text["run_inverse"], command=self.run_inverse)
        self.btn_inverse.grid(row=4, column=0, columnspan=2, pady=6, sticky="w")
        self.btn_misclosure = ttk.Button(frame, text=self.text["run_misclosure"], command=self.run_misclosure)
        self.btn_misclosure.grid(row=4, column=2, columnspan=2, pady=6, sticky="w")

        self.fwd_title = ttk.Label(frame, text=self.text["forward_title"], font=("Segoe UI", 10, "bold"))
        self.fwd_title.grid(row=5, column=0, columnspan=4, sticky="w", pady=(12, 6))

        ttk.Label(frame, text=self.text["b1"]).grid(row=6, column=0, sticky="w")
        self.fwd_b1 = ttk.Entry(frame, width=12)
        self.fwd_b1.grid(row=6, column=1, padx=6, pady=2)
        ttk.Label(frame, text=self.text["l1"]).grid(row=6, column=2, sticky="w")
        self.fwd_l1 = ttk.Entry(frame, width=12)
        self.fwd_l1.grid(row=6, column=3, padx=6, pady=2)
        ttk.Label(frame, text=self.text["az"]).grid(row=7, column=0, sticky="w")
        self.fwd_az = ttk.Entry(frame, width=12)
        self.fwd_az.grid(row=7, column=1, padx=6, pady=2)
        ttk.Label(frame, text=self.text["dist"]).grid(row=7, column=2, sticky="w")
        self.fwd_s = ttk.Entry(frame, width=12)
        self.fwd_s.grid(row=7, column=3, padx=6, pady=2)

        self.btn_forward = ttk.Button(frame, text=self.text["run_forward"], command=self.run_forward)
        self.btn_forward.grid(row=8, column=0, columnspan=2, pady=6, sticky="w")
        self.btn_reduction = ttk.Button(frame, text=self.text["run_reduction"], command=self.run_reduction)
        self.btn_reduction.grid(row=8, column=2, columnspan=2, pady=6, sticky="w")

        self.geo_output = tk.Text(frame, height=18, wrap="word")
        self.geo_output.grid(row=9, column=0, columnspan=4, sticky="nsew", pady=8)
        self.lbl_history = ttk.Label(frame, text=self.text["history"])
        self.lbl_history.grid(row=10, column=0, sticky="w")
        self.history_text = tk.Text(frame, height=6, wrap="word")
        self.history_text.grid(row=11, column=0, columnspan=4, sticky="nsew", pady=4)
        frame.rowconfigure(9, weight=1)
        frame.rowconfigure(11, weight=1)
        for i in range(4):
            frame.columnconfigure(i, weight=1)

    def build_special_tab(self) -> None:
        grid = ttk.Frame(self.tab_special)
        grid.pack(fill=tk.X)

        self.special_labels = {}
        keys = ["b1", "l1", "h1", "b2", "l2", "h2"]
        self.special_entries = {}
        for i, key in enumerate(keys):
            lbl = ttk.Label(grid, text=self.text[key])
            lbl.grid(row=i // 3, column=(i % 3) * 2, sticky="w", padx=2, pady=4)
            ent = ttk.Entry(grid, width=14)
            ent.grid(row=i // 3, column=(i % 3) * 2 + 1, sticky="w", padx=4, pady=4)
            self.special_labels[key] = lbl
            self.special_entries[key] = ent

        self.btn_special = ttk.Button(self.tab_special, text=self.text["special_run"], command=self.run_special)
        self.btn_special.pack(anchor="w", pady=6)
        self.btn_special_pdf = ttk.Button(self.tab_special, text=self.text["export_pdf"], command=self.export_special_pdf)
        self.btn_special_pdf.pack(anchor="w", pady=2)

        self.special_out = tk.Text(self.tab_special, height=28, wrap="word")
        self.special_out.pack(fill=tk.BOTH, expand=True)

    def import_txt(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if not path:
            return
        with open(path, "r", encoding="utf-8") as f:
            data = f.read()
        self.input_text.delete("1.0", tk.END)
        self.input_text.insert("1.0", data)

    def _collect_output_points(self) -> List[Tuple[float, float, float | None]]:
        pts: List[Tuple[float, float, float | None]] = []
        for line in self.output_text.get("1.0", tk.END).splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                vals = parse_row(line)
                if len(vals) >= 3:
                    pts.append((vals[0], vals[1], vals[2]))
                elif len(vals) >= 2:
                    pts.append((vals[0], vals[1], None))
            except Exception:
                continue
        return pts

    def export_txt(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if not path:
            return
        data = self.output_text.get("1.0", tk.END)
        with open(path, "w", encoding="utf-8") as f:
            f.write(data)
        self.log(f"TXT exported: {path}")

    def export_kml(self) -> None:
        pts = self._collect_output_points()
        if not pts:
            messagebox.showwarning("KML", "No points in output.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".kml", filetypes=[("KML", "*.kml")])
        if not path:
            return
        with open(path, "w", encoding="utf-8") as f:
            f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
            f.write('<kml xmlns="http://www.opengis.net/kml/2.2"><Document>\n')
            for i, (x, y, z) in enumerate(pts, 1):
                zz = 0.0 if z is None else z
                f.write(f"<Placemark><name>P{i}</name><Point><coordinates>{x},{y},{zz}</coordinates></Point></Placemark>\n")
            f.write("</Document></kml>\n")
        self.log(f"KML exported: {path}")

    def export_gpx(self) -> None:
        pts = self._collect_output_points()
        if not pts:
            messagebox.showwarning("GPX", "No points in output.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".gpx", filetypes=[("GPX", "*.gpx")])
        if not path:
            return
        with open(path, "w", encoding="utf-8") as f:
            f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
            f.write('<gpx version="1.1" creator="GeoMate">\n')
            for i, (x, y, z) in enumerate(pts, 1):
                # assumes output contains lon,lat,(h)
                if z is None:
                    f.write(f'<wpt lon="{x}" lat="{y}"><name>P{i}</name></wpt>\n')
                else:
                    f.write(f'<wpt lon="{x}" lat="{y}"><ele>{z}</ele><name>P{i}</name></wpt>\n')
            f.write("</gpx>\n")
        self.log(f"GPX exported: {path}")

    def export_dxf(self) -> None:
        pts = self._collect_output_points()
        if not pts:
            messagebox.showwarning("DXF", "No points in output.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".dxf", filetypes=[("DXF", "*.dxf")])
        if not path:
            return
        with open(path, "w", encoding="utf-8") as f:
            f.write("0\nSECTION\n2\nENTITIES\n")
            for x, y, z in pts:
                zz = 0.0 if z is None else z
                f.write("0\nPOINT\n8\n0\n")
                f.write(f"10\n{x}\n20\n{y}\n30\n{zz}\n")
            f.write("0\nENDSEC\n0\nEOF\n")
        self.log(f"DXF exported: {path}")

    def preview_map(self) -> None:
        if not HAS_MPL:
            messagebox.showerror("Map", self.text["mpl_missing"])
            return
        pts = self._collect_output_points()
        if not pts:
            messagebox.showwarning("Map", "No points in output.")
            return
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        fig, ax = plt.subplots(figsize=(7, 5))
        ax.plot(xs, ys, "o-", linewidth=1)
        ax.set_title("GeoMate Preview")
        ax.set_xlabel("X / Lon")
        ax.set_ylabel("Y / Lat")
        ax.grid(True, alpha=0.3)
        temp = os.path.join(tempfile.gettempdir(), "geomate_preview.png")
        fig.savefig(temp, dpi=140, bbox_inches="tight")
        plt.close(fig)
        try:
            webbrowser.open(temp)
        except Exception:
            pass
        self.log("Preview map generated")

    def import_table(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return

        rows: List[str] = []
        if path.lower().endswith(".csv"):
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                for r in reader:
                    if r:
                        rows.append(" ".join(c.strip() for c in r if c.strip()))
        elif path.lower().endswith(".xlsx"):
            if not HAS_OPENPYXL:
                messagebox.showerror("XLSX", self.text["xlsx_missing"])
                return
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                vals = [str(v).strip() for v in r if v is not None and str(v).strip() != ""]
                if vals:
                    rows.append(" ".join(vals))
        else:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                rows = [ln.strip() for ln in f if ln.strip()]

        self.input_text.delete("1.0", tk.END)
        self.input_text.insert("1.0", "\n".join(rows))

    def export_table(self) -> None:
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        lines = [ln.strip() for ln in self.output_text.get("1.0", tk.END).splitlines() if ln.strip()]
        if path.lower().endswith(".xlsx"):
            if not HAS_OPENPYXL:
                messagebox.showerror("XLSX", self.text["xlsx_missing"])
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            for i, ln in enumerate(lines, start=1):
                vals = [v for v in ln.replace("\t", " ").split(" ") if v]
                for j, v in enumerate(vals, start=1):
                    ws.cell(i, j, v)
            wb.save(path)
        else:
            with open(path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                for ln in lines:
                    vals = [v for v in ln.replace("\t", " ").split(" ") if v]
                    writer.writerow(vals)

    def save_project(self) -> None:
        path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("GeoMate project", "*.json"), ("All files", "*.*")])
        if not path:
            return
        payload = {
            "lang": self.lang_var.get(),
            "in_crs": self.entry_in_crs.get().strip(),
            "out_crs": self.entry_out_crs.get().strip(),
            "order": self.order_var.get(),
            "datum": self.datum_var.get(),
            "height_model": self.height_model_var.get(),
            "geoid_n": self.entry_geoid_n.get().strip(),
            "apply_height": self.apply_height_var.get(),
            "input_text": self.input_text.get("1.0", tk.END),
            "output_text": self.output_text.get("1.0", tk.END),
            "special": {k: self.special_entries[k].get().strip() for k in self.special_entries},
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    def load_project(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("GeoMate project", "*.json"), ("All files", "*.*")])
        if not path:
            return
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        self.lang_var.set(payload.get("lang", "RU"))
        self.on_language_change()
        self.entry_in_crs.delete(0, tk.END)
        self.entry_in_crs.insert(0, payload.get("in_crs", "EPSG:4326"))
        self.entry_out_crs.delete(0, tk.END)
        self.entry_out_crs.insert(0, payload.get("out_crs", "EPSG:3857"))
        self.order_var.set(payload.get("order", self.text["latlon"]))
        self.datum_var.set(payload.get("datum", "WGS84"))
        self.height_model_var.set(payload.get("height_model", "Ellipsoidal"))
        self.entry_geoid_n.delete(0, tk.END)
        self.entry_geoid_n.insert(0, payload.get("geoid_n", "0.0"))
        self.apply_height_var.set(bool(payload.get("apply_height", False)))
        self.input_text.delete("1.0", tk.END)
        self.input_text.insert("1.0", payload.get("input_text", ""))
        self.output_text.delete("1.0", tk.END)
        self.output_text.insert("1.0", payload.get("output_text", ""))
        special = payload.get("special", {})
        for k, ent in self.special_entries.items():
            ent.delete(0, tk.END)
            ent.insert(0, str(special.get(k, "")))

    def _decode_order(self) -> Tuple[bool, bool]:
        # returns (latlon_input, has_z)
        order = self.order_var.get()
        if order in (self.text["latlon"],):
            return True, False
        if order in (self.text["latlonh"],):
            return True, True
        if order in (self.text["xyz"],):
            return False, True
        return False, False

    def run_transform(self) -> None:
        if not HAS_PYPROJ:
            messagebox.showerror("pyproj", self.text["pyproj_missing"])
            return

        try:
            in_crs_raw = self.entry_in_crs.get().strip()
            out_crs_raw = self.entry_out_crs.get().strip()

            preset = self.datum_var.get()
            if preset == "WGS84":
                if in_crs_raw == "":
                    in_crs_raw = "EPSG:4326"
            elif preset == "SK-42":
                if in_crs_raw == "":
                    in_crs_raw = "EPSG:4284"
            elif preset == "SK-95":
                if in_crs_raw == "":
                    in_crs_raw = "EPSG:4179"
            elif preset == "PZ-90.11":
                if in_crs_raw == "":
                    in_crs_raw = "EPSG:7683"

            in_crs = CRS.from_user_input(in_crs_raw)
            out_crs = CRS.from_user_input(out_crs_raw)
            transformer = Transformer.from_crs(in_crs, out_crs, always_xy=True)
        except Exception as e:
            messagebox.showerror("CRS", str(e))
            return

        latlon_input, has_z = self._decode_order()
        try:
            geoid_n = float(self.entry_geoid_n.get().replace(",", "."))
        except Exception:
            geoid_n = 0.0
        apply_h = self.apply_height_var.get()
        lines = self.input_text.get("1.0", tk.END).splitlines()
        out = io.StringIO()
        writer = csv.writer(out, delimiter="\t", lineterminator="\n")

        for line in lines:
            if not line.strip():
                continue
            try:
                vals = parse_row(line)
                if has_z and len(vals) < 3:
                    raise ValueError("need 3 values")
                if (not has_z) and len(vals) < 2:
                    raise ValueError("need 2 values")

                if has_z:
                    a, b, z = vals[0], vals[1], vals[2]
                else:
                    a, b = vals[0], vals[1]
                    z = None

                if latlon_input:
                    lon, lat = b, a
                else:
                    lon, lat = a, b

                if z is None:
                    x2, y2 = transformer.transform(lon, lat)
                    writer.writerow([f"{x2:.6f}", f"{y2:.6f}"])
                else:
                    x2, y2, z2 = transformer.transform(lon, lat, z)
                    if apply_h and self.height_model_var.get() != "Ellipsoidal":
                        z2 = z2 - geoid_n
                    writer.writerow([f"{x2:.6f}", f"{y2:.6f}", f"{z2:.6f}"])
            except Exception as e:
                writer.writerow([f"{self.text['err_parse']} {line} ({e})"])

        self.output_text.delete("1.0", tk.END)
        self.output_text.insert("1.0", out.getvalue())
        self.log("CRS transform finished")

    def run_inverse(self) -> None:
        try:
            use_dms = self.angle_mode_var.get() == self.text["dms_mode_value"]
            b1 = angle_from_user(self.inv_b1.get(), use_dms)
            l1 = angle_from_user(self.inv_l1.get(), use_dms)
            b2 = angle_from_user(self.inv_b2.get(), use_dms)
            l2 = angle_from_user(self.inv_l2.get(), use_dms)
            az, dist = core.vincenty_inverse(b1, l1, b2, l2)
            txt = f"A12 = {az:.8f} deg\nS12 = {dist:.4f} m\n"
            self.geo_output.delete("1.0", tk.END)
            self.geo_output.insert("1.0", txt)
            self.log(f"Inverse solved: A12={az:.6f}, S12={dist:.3f}")
        except Exception as e:
            messagebox.showerror("Inverse", str(e))

    def run_forward(self) -> None:
        try:
            use_dms = self.angle_mode_var.get() == self.text["dms_mode_value"]
            b1 = angle_from_user(self.fwd_b1.get(), use_dms)
            l1 = angle_from_user(self.fwd_l1.get(), use_dms)
            az = angle_from_user(self.fwd_az.get(), use_dms)
            dist = float(self.fwd_s.get().replace(",", "."))
            b2, l2 = core.vincenty_forward(b1, l1, az, dist)
            txt = f"B2 = {b2:.8f} deg\nL2 = {l2:.8f} deg\n"
            self.geo_output.delete("1.0", tk.END)
            self.geo_output.insert("1.0", txt)
            self.log(f"Forward solved: B2={b2:.8f}, L2={l2:.8f}")
        except Exception as e:
            messagebox.showerror("Forward", str(e))

    def run_reduction(self) -> None:
        try:
            if not HAS_PYPROJ:
                raise RuntimeError(self.text["pyproj_missing"])
            b1 = angle_from_user(self.fwd_b1.get(), self.angle_mode_var.get() == self.text["dms_mode_value"])
            l1 = angle_from_user(self.fwd_l1.get(), self.angle_mode_var.get() == self.text["dms_mode_value"])
            s = float(self.fwd_s.get().replace(",", "."))
            in_crs = CRS.from_user_input(self.entry_in_crs.get().strip() or "EPSG:4326")
            if in_crs.is_geographic:
                proj = Transformer.from_crs(in_crs, CRS.from_user_input(self.entry_out_crs.get().strip() or "EPSG:3857"), always_xy=True)
                # fallback: no direct factors for transformer pipeline -> approximate k=1
                k = 1.0
            else:
                k = 1.0
            s_plane = s * k
            self.geo_output.delete("1.0", tk.END)
            self.geo_output.insert("1.0", f"k = {k:.8f}\nS = {s:.4f} m\nS' = k*S = {s_plane:.4f} m\n")
            self.log(f"Reduction computed: k={k:.6f}")
        except Exception as e:
            messagebox.showerror("Reduction", str(e))

    def run_misclosure(self) -> None:
        """
        Simple traverse misclosure from output points sequence.
        Uses first and last point closure vectors.
        """
        try:
            pts = self._collect_output_points()
            if len(pts) < 2:
                raise RuntimeError("Need at least 2 points in output.")
            x0, y0, _ = pts[0]
            xn, yn, _ = pts[-1]
            fx = xn - x0
            fy = yn - y0
            f = math.hypot(fx, fy)
            self.geo_output.delete("1.0", tk.END)
            self.geo_output.insert("1.0", f"Linear misclosure:\nfx={fx:.4f}\nfy={fy:.4f}\nf={f:.4f}\n")
            self.log(f"Misclosure computed: f={f:.4f}")
        except Exception as e:
            messagebox.showerror("Misclosure", str(e))

    def run_special(self) -> None:
        try:
            b1 = float(self.special_entries["b1"].get().replace(",", "."))
            l1 = float(self.special_entries["l1"].get().replace(",", "."))
            h1 = float(self.special_entries["h1"].get().replace(",", "."))
            b2 = float(self.special_entries["b2"].get().replace(",", "."))
            l2 = float(self.special_entries["l2"].get().replace(",", "."))
            h2 = float(self.special_entries["h2"].get().replace(",", "."))

            zone1 = core.get_zone_number(l1)
            zone2 = zone1 + 1
            x1, y1 = core.geodetic_to_gauss(b1, l1, zone1)
            b1b, l1b = core.gauss_to_geodetic(x1, y1, zone1)
            x1g, y1g = core.gost_zone_transform(x1, y1, zone1, zone2)
            x1t, y1t = core.thompson_algorithm(x1, y1, zone1, zone2)
            x1k, y1k = core.krueger_algorithm(x1, y1, zone1, zone2)
            x1r, y1r = core.gerasimenko_algorithm(x1, y1, zone1, zone2)
            az, dist = core.vincenty_inverse(b1, l1, b2, l2)
            b2f, l2f = core.vincenty_forward(b1, l1, az, dist)
            x2, y2 = core.geodetic_to_gauss(b2f, l2f, zone2)
            b2c, l2c = core.gauss_to_geodetic(x2, y2, zone2)
            b95, l95, h95 = core.sk42_to_sk95(b2, l2, h2)

            report = []
            report.append(f"zone1={zone1}, zone2={zone2}")
            report.append(f"x1={x1:.3f}  y1={y1:.3f}")
            report.append(f"control B1*={b1b:.10f} L1*={l1b:.10f}")
            report.append(f"x1' GOST={x1g:.3f} {y1g:.3f}")
            report.append(f"x1' Thompson={x1t:.3f} {y1t:.3f}")
            report.append(f"x1' Krueger={x1k:.3f} {y1k:.3f}")
            report.append(f"x1' Gerasimenko={x1r:.3f} {y1r:.3f}")
            report.append(f"A12={az:.8f}  S12={dist:.3f}")
            report.append(f"x2={x2:.3f}  y2={y2:.3f}")
            report.append(f"control B2*={b2c:.10f} L2*={l2c:.10f}")
            report.append(f"SK95: B2={b95:.10f} L2={l95:.10f} H2={h95:.3f}")
            report.append(f"H1={h1:.3f}, H2={h2:.3f}")

            self.special_out.delete("1.0", tk.END)
            self.special_out.insert("1.0", "\n".join(report))
            self.log("Special workflow completed")
        except Exception as e:
            messagebox.showerror("Special", str(e))

    def export_special_pdf(self) -> None:
        if not HAS_MPL:
            messagebox.showerror("PDF", self.text["mpl_missing"])
            return
        try:
            # Build a tiny scheme from point 1 and 2 in neighboring zone
            b1 = float(self.special_entries["b1"].get().replace(",", "."))
            l1 = float(self.special_entries["l1"].get().replace(",", "."))
            b2 = float(self.special_entries["b2"].get().replace(",", "."))
            l2 = float(self.special_entries["l2"].get().replace(",", "."))
            zone1 = core.get_zone_number(l1)
            zone2 = zone1 + 1
            x1, y1 = core.geodetic_to_gauss(b1, l1, zone2)
            x2, y2 = core.geodetic_to_gauss(b2, l2, zone2)
            path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
            if not path:
                return
            fig, ax = plt.subplots(figsize=(8, 6))
            ax.plot([x1, x2], [y1, y2], "o-", lw=1.8)
            ax.annotate("P1", (x1, y1))
            ax.annotate("P2", (x2, y2))
            ax.set_title("GeoMate Special Workflow Sketch")
            ax.set_xlabel("X (m)")
            ax.set_ylabel("Y (m)")
            ax.grid(True, alpha=0.3)
            fig.savefig(path, bbox_inches="tight")
            plt.close(fig)
            self.log(f"PDF scheme exported: {path}")
        except Exception as e:
            messagebox.showerror("PDF", str(e))


def run() -> None:
    root = tk.Tk()
    GeoMateApp(root)
    root.mainloop()


if __name__ == "__main__":
    run()
